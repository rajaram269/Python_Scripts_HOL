import pandas as pd
from openpyxl import load_workbook
from sklearn.ensemble import RandomForestClassifier
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.pipeline import Pipeline
import pickle
import os

# Training Data for ML Model
training_data = [
    ("inventory,inv, Stock, Warehouse,store available", "SOH"),
    ("Invoice, Customer, Amount, mrp, mrp value", "Sales"),
    ("SKU, Stock, Warehouse, inventory, inv, Invoice, Customer, Amount", "SOH_Sales"),
    ("Random, Unrelated, Headers", "Unknown"),
]

# Train and Save the Classifier
def train_sheet_classifier(training_data):
    """
    Train a classifier to identify sheet types based on column headers.
    """
    X_train = [item[0] for item in training_data]
    y_train = [item[1] for item in training_data]
    
    pipeline = Pipeline([
        ("vectorizer", TfidfVectorizer()),
        ("classifier", RandomForestClassifier())
    ])
    
    pipeline.fit(X_train, y_train)
    
    with open("sheet_classifier.pkl", "wb") as f:
        pickle.dump(pipeline, f)
    print("Model trained and saved as 'sheet_classifier.pkl'")

# Load the Classifier
def load_sheet_classifier():
    with open("sheet_classifier.pkl", "rb") as f:
        #print (pickle.load(f))
        return pickle.load(f)

# Field Mapping Dictionary
field_mapping = {
    "SKU": ["SKU", "Product ID","Style Code"],
    "Stock": ["Stock", "Inventory", "qty available"],
    "Warehouse": ["Warehouse", "Depot","location"],
    "Invoice_Number": ["Invoice"],
    "Customer_Name": ["Customer"],
    "Sales_Amount": ["Amount"]
}

def map_fields(headers):
    """
    Map raw headers to standard field names.
    """
    mapped_fields = {}
    for field, synonyms in field_mapping.items():
        for synonym in synonyms:
            if synonym in headers:
                mapped_fields[synonym] = field
    return mapped_fields

def clean_data(df):
    """
    Clean erratic or unhygienic data in a DataFrame.
    - Identifies the header row as the first row with all columns filled.
    - Removes rows above the header row.
    - Cleans the dataset by removing empty rows/columns and duplicates.
    """
    header_row = None
    for idx, row in df.iterrows():
        if row.notna().all():
            header_row = idx
            break

    if header_row is None:
        raise ValueError("No valid header row found. All rows have missing columns.")

    df = df.iloc[header_row:].reset_index(drop=True)
    df.columns = df.iloc[0].str.strip().str.lower()
    df = df[1:]
    df.dropna(how="all", inplace=True)
    df.dropna(how="all", axis=1, inplace=True)
    df.drop_duplicates(inplace=True)
    df.fillna(method="ffill", inplace=True)
    return df

# Main File Processor
def process_file(file_path, output_path):
    """
    Process a CSV or Excel file to classify, map, and clean sheets.
    """
    # Detect file type
    file_extension = os.path.splitext(file_path)[1].lower()
    classifier = load_sheet_classifier()
    results = {}

    if file_extension == ".xlsx":
        wb = load_workbook(file_path)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            df = pd.DataFrame(ws.values)
            if df.empty:
                continue
            process_dataframe(df, sheet, results, classifier)
    elif file_extension == ".csv":
        df = pd.read_csv(file_path)
        process_dataframe(df, "Sheet1", results, classifier)
    else:
        raise ValueError("Unsupported file format. Only .xlsx and .csv are supported.")

    # Save results
    save_results(results, output_path, file_extension)

def process_dataframe(df, sheet_name, results, classifier):
    """
    Process an individual DataFrame by classifying, mapping, and cleaning it.
    """
    #print(classifier)
    try:
        cleaned_df = clean_data(df)
        #print(cleaned_df.columns)
        #results[sheet_name] = (sheet_type, cleaned_df)
    except Exception as e:
        print(f"Error cleaning data in sheet '{sheet_name}': {e}")

    headers = cleaned_df.columns.str.lower().str.strip()
    headers=[str(h) for h in headers]
    headers_string = ", ".join(headers)
    print(headers_string)

    try:
        #print(1)
      sheet_type = classifier.predict([headers_string])[0]
    #print('sheet type is:' & sheet_type)
    except Exception as e:
        print(f"Error classifying sheet '{sheet_name}': {e}")
        sheet_type = "Unknown"

    print(sheet_type)
    mapped_fields = map_fields(headers)
    print(mapped_fields)
    try:
        cleaned_df.rename(columns=mapped_fields, inplace=True)
    except Exception as e:
        print(f"Error mapping fields in sheet '{sheet_name}': {e}")

    try:
        #cleaned_df1 = clean_data(cleaned_df)
        results[sheet_name] = (sheet_type, cleaned_df)
    except Exception as e:
        print(f"Error cleaning data in sheet '{sheet_name}': {e}")

def save_results(results, output_path, file_extension):
    """
    Save processed results to a new file.
    """
    if file_extension == ".xlsx":
        with pd.ExcelWriter(output_path) as writer:
            for sheet_name, (sheet_type, df) in results.items():
                df.to_excel(writer, sheet_name=f"{sheet_type}_{sheet_name}", index=False)
    elif file_extension == ".csv":
        for sheet_name, (sheet_type, df) in results.items():
            output_csv_path = output_path.replace(".csv", f"_{sheet_type}_{sheet_name}.csv")
            df.to_csv(output_csv_path, index=False)

# Train the model (Run this once to prepare the model)
train_sheet_classifier(training_data)

# Process a file
file_path = "E:/Extract/Enrich.xlsx"  # Change to "input_file.xlsx" for Excel
output_path = "E:/Extract/processed_file_Enrich.xlsx"  # Change to "processed_file.xlsx" for Excel
process_file(file_path, output_path)
