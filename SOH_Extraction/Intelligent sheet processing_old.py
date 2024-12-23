import pandas as pd
import os
import sklearn
from openpyxl import load_workbook
from sklearn.ensemble import RandomForestClassifier
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.pipeline import Pipeline
import pickle

# Expanded Training Data for ML Model
training_data = [
    ("SKU, Stock, Warehouse,qty available", "SOH"),
    ("Invoice, Customer, Amount", "Sales"),
    ("SKU, Stock, Warehouse, Invoice, Customer, Amount", "SOH_Sales"),
    ("Random, Unrelated, Headers", "Unknown"),
    ("Product, Quantity, Location", "Inventory"),
    ("Date, Sale, Price", "Sales"),
]

# Train and Save the Classifier
def train_sheet_classifier(training_data):
    """
    Train a classifier to identify sheet types based on column headers.
    """
    try:
        X_train = [item[0] for item in training_data]
        y_train = [item[1] for item in training_data]
        
        pipeline = Pipeline([
            ("vectorizer", TfidfVectorizer()),
            ("classifier", RandomForestClassifier(random_state=42))
        ])
        
        pipeline.fit(X_train, y_train)
        
        # Ensure directory exists
        os.makedirs(os.path.dirname("models/"), exist_ok=True)
        
        with open("models/sheet_classifier.pkl", "wb") as f:
            pickle.dump(pipeline, f)
        print("Model trained and saved as 'models/sheet_classifier.pkl'")
    except Exception as e:
        print(f"Error training classifier: {e}")

# Load the Classifier with Error Handling
def load_sheet_classifier():
    try:
        with open("models/sheet_classifier.pkl", "rb") as f:
            return pickle.load(f)
    except FileNotFoundError:
        print("Classifier model not found. Training a new model.")
        train_sheet_classifier(training_data)
        return load_sheet_classifier()
    except Exception as e:
        print(f"Error loading classifier: {e}")
        raise

# Field Mapping Dictionary (Expanded)
field_mapping = {
    "Stock Keeping Unit": ["SKU", "Product ID", "Product Code"],
    "Stock": ["Stock", "Inventory", "Quantity", "Units", "qty available"],
    "Warehouse": ["Warehouse", "Depot", "Location", "Storage"],
    "Invoice Number": ["Invoice", "Bill Number", "Transaction ID"],
    "Customer Name": ["Customer", "Client", "Buyer"],
    "Sales Amount": ["Amount", "Price", "Total", "Value"]
}

def map_fields(headers):
    """
    Map raw headers to standard field names.
    """
    mapped_fields = {}
    for field, synonyms in field_mapping.items():
        for synonym in synonyms:
            matching_headers = [h for h in headers if synonym.lower() in h.lower()]
            if matching_headers:
                mapped_fields[matching_headers[0]] = field
    return mapped_fields

def clean_data(df):
    """
    Clean erratic or unhygienic data in a DataFrame.
    """
    # Try multiple approaches to find header row
    for header_guess in range(min(5, len(df))):
        try:
            # Check if row has unique values and all columns filled
            headers = df.iloc[header_guess].str.strip().str.lower()
            if len(set(headers)) == len(headers) and headers.notna().all():
                df = df.iloc[header_guess:].reset_index(drop=True)
                df.columns = headers
                df = df[1:]
                break
        except Exception:
            continue
    
    df.dropna(how="all", inplace=True)
    df.dropna(how="all", axis=1, inplace=True)
    df.drop_duplicates(inplace=True)
    
    # Fill missing values more robustly
    for col in df.columns:
        if df[col].dtype in ['object']:
            df[col].fillna('Unknown', inplace=True)
        else:
            df[col].fillna(df[col].median(), inplace=True)
    
    return df

# Main File Processor with Enhanced Error Handling
def process_file(file_path, output_path=None):
    """
    Process a CSV or Excel file to classify, map, and clean sheets.
    """
    # Validate file path
    if not os.path.exists(file_path):
        print(f"Error: File {file_path} not found.")
        return None

    # Default output path if not provided
    if output_path is None:
        base, ext = os.path.splitext(file_path)
        output_path = f"{base}_processed{ext}"

    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    try:
        classifier = load_sheet_classifier()
        results = {}

        file_extension = os.path.splitext(file_path)[1].lower()

        if file_extension == ".xlsx":
            wb = load_workbook(file_path, read_only=True)
            for sheet in wb.sheetnames:
                df = pd.DataFrame(wb[sheet].values)
                if not df.empty:
                    process_dataframe(df, sheet, results, classifier)
        elif file_extension == ".csv":
            df = pd.read_csv(file_path, encoding='utf-8', low_memory=False)
            process_dataframe(df, "Sheet1", results, classifier)
        else:
            print(f"Unsupported file format: {file_extension}. Only .xlsx and .csv are supported.")
            return None

        # Save results
        save_results(results, output_path, file_extension)
        return results

    except Exception as e:
        print(f"Error processing file {file_path}: {e}")
        return None

def process_dataframe(df, sheet_name, results, classifier):
    """
    Process an individual DataFrame by classifying, mapping, and cleaning it.
    """
    try:
        # Convert first row to lowercase and strip
        headers = df.iloc[0].astype(str).str.lower().str.strip()
        headers_string = ", ".join(headers.dropna())

        # Predict sheet type
        sheet_type = classifier.predict([headers_string])[0]
        
        # Map and rename fields
        mapped_fields = map_fields(headers)
        df.rename(columns=mapped_fields, inplace=True)

        # Clean data
        cleaned_df = clean_data(df)
        results[sheet_name] = (sheet_type, cleaned_df)

    except Exception as e:
        print(f"Error processing sheet '{sheet_name}': {e}")

def save_results(results, output_path, file_extension):
    """
    Save processed results to a new file.
    """
    try:
        if file_extension == ".xlsx":
            with pd.ExcelWriter(output_path) as writer:
                for sheet_name, (sheet_type, df) in results.items():
                    df.to_excel(writer, sheet_name=f"{sheet_type}_{sheet_name}", index=False)
        elif file_extension == ".csv":
            for sheet_name, (sheet_type, df) in results.items():
                output_csv_path = output_path.replace(file_extension, f"_{sheet_type}_{sheet_name}{file_extension}")
                df.to_csv(output_csv_path, index=False)
        
        print(f"Results saved to {output_path}")
    except Exception as e:
        print(f"Error saving results: {e}")

# Train the model (uncomment if model doesn't exist)
# train_sheet_classifier(training_data)

# Example usage
def main():
    # Make these relative or full paths to your actual files
    input_files = [
        "E:\Extract\SSL.xlsx",  # Replace with your CSV file
      #  "E:\Extract\sample_input.xlsx"  # Replace with your Excel file
    ]
    
    for file_path in input_files:
        if os.path.exists(file_path):
            print(f"Processing {file_path}")
            process_file(file_path)
        else:
            print(f"File not found: {file_path}")

if __name__ == "__main__":
    main()